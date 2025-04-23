import os
import pandas as pd
import numpy as np
import sqlite3
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule


# Obtener el directorio del perfil del usuario actual:
user_profile = os.environ.get("USERPROFILE")

# Construir la ruta a la carpeta de OneDrive de la empresa. Por ejemplo:
base_path = os.path.join(user_profile, "OneDrive - Telefonica", "Dalia Paola Rodriguez Cruz's files - TOA_proceso")


def procesar_archivos_tickets(carpeta, tabla, conexion, id):
    """
    Procesa los archivos Excel en la carpeta especificada y actualiza la base de datos.

    Este proceso realiza lo siguiente:
      - Obtiene la lista de archivos Excel nuevos (no procesados) en la carpeta.
      - Combina los datos de dichos archivos en un único DataFrame.
      - Elimina la columna 'Mes' en caso de existir, ya que no es requerida.
      - Actualiza la base de datos con el DataFrame combinado.
      - Marca los archivos procesados para evitar reprocesamientos futuros.

    Args:
        carpeta (str): Ruta de la carpeta que contiene los archivos.
        tabla (str): Nombre de la tabla en la base de datos donde se actualizarán los datos.
        conexion (sqlite3.Connection): Conexión activa a la base de datos SQLite.
        id (str): Nombre de la columna identificadora para eliminar duplicados.
    """
    # Se obtienen los archivos Excel a procesar
    archivos = obtener_archivos_excel(carpeta)
    
    if not archivos:
        print("\tNo se encontraron archivos nuevos para procesar.")
        return

    # Se combinan los datos de los archivos en un solo DataFrame
    df_final = combinar_datos_archivos(carpeta, archivos)

    # Si hay datos combinados y el DataFrame no está vacío
    if df_final is not None and not df_final.empty:
        # Si existe la columna 'Mes', se elimina ya que no es necesaria para el procesamiento
        if 'Mes' in df_final.columns:
            df_final.drop(columns=['Mes'], inplace=True)
        print("Ya se puede actualizar la base de datos.")
        actualizar_base_datos(conexion, tabla, df_final, id)
        marcar_archivos_procesados(carpeta, archivos)
    else:
        print("\tNo hay datos nuevos para actualizar.")


def obtener_archivos_excel(carpeta):
    """
    Obtiene una lista de archivos Excel en la carpeta que aún no han sido procesados.

    Se excluyen aquellos archivos que:
      - Tengan extensión '_procesado.xlsx' o '_procesado.xls' para evitar reprocesamiento.
      - Tengan nombre 'deskto.xlsx' (ignora archivos no relevantes).
      - No sean archivos (por ejemplo, directorios).

    Además, se intenta ordenar los archivos que contienen fecha en el nombre.

    Args:
        carpeta (str): Ruta de la carpeta que contiene los archivos.

    Returns:
        list: Lista de nombres de archivos a procesar.
    """
    archivos = [f for f in os.listdir(carpeta) 
                if f.endswith(('.xlsx', '.xls')) and 
                not f.endswith('_procesado.xlsx') and 
                not f.endswith('_procesado.xls') and 
                f.lower() != 'deskto.xlsx' and 
                os.path.isfile(os.path.join(carpeta, f))]

    def ordenar_por_fecha(archivo):
        """
        Función interna para extraer la fecha del nombre del archivo.

        Se asume que el nombre del archivo comienza con el día y mes (separados por puntos).
        La lógica asigna:
          - Año 2024 si el mes es mayor o igual a 8.
          - Año 2025 en caso contrario.
          
        Nota: Esta lógica funcionará correctamente hasta junio de 2025. A partir de julio de 2025,
        se deberá actualizar para reflejar el año real.

        Args:
            archivo (str): Nombre del archivo.

        Returns:
            tuple or None: Tupla (año, mes, día) si se pudo extraer la fecha, o None en caso contrario.
        """
        try:
            dia, mes = map(int, archivo.split('.')[0:2])
            año = 2024 if mes >= 8 else 2025
            return (año, mes, dia)
        except (ValueError, IndexError):
            return None  # Si falla, se ordenará alfabéticamente

    # Se separan los archivos que tienen fecha en el nombre de los que no
    archivos_fecha = [archivo for archivo in archivos if ordenar_por_fecha(archivo) is not None]
    archivos_sin_fecha = [archivo for archivo in archivos if ordenar_por_fecha(archivo) is None]
    
    archivos_fecha.sort(key=ordenar_por_fecha)
    archivos_sin_fecha.sort()
    
    # Se devuelve la lista combinada: primero los archivos sin fecha, luego los ordenados por fecha
    return archivos_sin_fecha + archivos_fecha


def combinar_datos_archivos(carpeta, archivos):
    """
    Combina los datos de los archivos Excel en un solo DataFrame.

    Se definen las columnas requeridas para el proceso y se valida que cada archivo
    contenga todas estas columnas. Dependiendo del contenido de la carpeta (ej. "TOA" o "PR"),
    se selecciona la hoja adecuada o se renombra la columna correspondiente.

    Args:
        carpeta (str): Ruta de la carpeta que contiene los archivos.
        archivos (list): Lista de archivos Excel a combinar.

    Returns:
        pd.DataFrame: DataFrame combinado con los datos más recientes.
                      Devuelve None si no se encontraron datos válidos.
    """
    # Definición de las columnas que se requieren en los archivos TOA
    columnas_requeridas = ['Técnico', 'ID Recurso', 'Nro TOA', 'Subtipo de Actividad', 'Número de Petición', 'Fecha de Cita', 'SLA Inicio', 'SLA Fin', 'Localidad', 'Dirección', 
                            'Direccion Polar X', 'Direccion Polar Y', 'Nombre Cliente', 'Hora de asignación de actividad', 'Fecha de Registro de actividad TOA', 
                            'Notas', 'Código de Cliente', 'Fecha Hora de Cancelación', 'Empresa', 'Bucket Inicial', 'Usuario - Iniciado', 'Nombre Distrito', 'Sistema Origen', 
                            'ID del Ticket', 'Quiebres', 'Fecha de Inicio PINT', 'Inicio PR1', 'Fin PR1', 'Fin PR2', 'Inicio PR2', 'Fin PR3', 'Inicio PR3', 'Fin PR4', 'Inicio PR4', 
                            'Motivo PR1', 'Motivo PR2', 'Motivo PR3', 'Motivo PR4', 'Nombre Local', 'Tipo de local', 'Zona geográfica', 'Zona', 'Estado TOA']

    df_list = []
    for archivo in archivos:
        ruta_completa = os.path.join(carpeta, archivo)
        print(f"▶️Procesando archivo: {ruta_completa}")
        # Lógica para archivos provenientes de TOA
        if "TOA base" in carpeta:
            hojas = pd.ExcelFile(ruta_completa, engine="openpyxl").sheet_names
            # Se busca la hoja 'Sheet1' o 'Page 1'
            if 'Sheet1' in hojas or 'Page 1' in hojas:
                sheet = 'Sheet1' if 'Sheet1' in hojas else 'Page 1'
                df = pd.read_excel(ruta_completa, sheet_name=sheet, engine="openpyxl")
                # Verificar que el archivo contenga todas las columnas requeridas
                if all(col in df.columns for col in columnas_requeridas):
                    df['origen'] = archivo  # Se añade la columna 'origen' para identificar el archivo
                    df = df[columnas_requeridas]
                    df_list.append(df)
                else:
                    print(f"Advertencia: El archivo {ruta_completa} no contiene todas las columnas requeridas. Saltando este archivo.")
                    continue
            else:
                print(f"Advertencia: El archivo {ruta_completa} no contiene una hoja llamada 'Sheet1'. Saltando este archivo.")
            continue
        # Lógica para archivos provenientes de la carpeta que contiene "PR"
        elif "Autin PR" in carpeta:
            df = pd.read_excel(ruta_completa)
            if 'Order ID' in df.columns:
                df_list.append(df)
            else:
                print(f"Advertencia: El archivo {ruta_completa} no contiene la columna 'Order ID'. Saltando este archivo.")
                continue
        # Lógica para otros archivos
        else:
            df = pd.read_excel(ruta_completa)
            # Renombrar la columna 'Nro TOA' a 'Number_OS_SIOM'
            df.rename(columns={'Nro TOA': 'Number_OS_SIOM'}, inplace=True)
            if 'Task Id' in df.columns:
                df_list.append(df)
            else:
                print(f"Advertencia: El archivo {ruta_completa} no contiene la columna 'Task Id'. Saltando este archivo.")
                continue
        
    # Si se han leído archivos válidos, se concatenan y se procesan
    if df_list:
        df_concatenado = pd.concat(df_list, ignore_index=True)
        # Reemplazar espacios en los nombres de columna por guiones bajos
        df_concatenado.columns = df_concatenado.columns.str.replace(' ', '_')
        if 'Mes' in df_concatenado.columns:
            df_concatenado.drop(columns=['Mes'], inplace=True)
        # Eliminar duplicados según la columna identificadora disponible
        if 'Nro_TOA' in df_concatenado.columns:
            df_reciente = df_concatenado.drop_duplicates(subset='Nro_TOA', keep='last')
        elif 'Task_Id' in df_concatenado.columns:
            df_reciente = df_concatenado.drop_duplicates(subset='Task_Id', keep='last')
        elif 'Order_ID' in df_concatenado.columns:
            df_reciente = df_concatenado
        else:
            print("No se encontraron columnas 'Nro TOA' o 'Task Id' para eliminar duplicados.")
            return None
        return df_reciente
    else:
        return None
    

def convertir_fechas(df, nombre_columna):
    """
    Convierte una columna del DataFrame a formato datetime unificado (YYYY-MM-DD HH:MM:SS).
    Si se encuentra algún valor que no se pueda convertir, se almacena en una lista para mostrarlos en consola.

    Nota: La función utiliza la expresión "convertir_fecha(x) or valores_no_convertidos.append(x)"
    para capturar valores no convertibles. Aunque la lógica actual funciona, se sugiere en futuras mejoras
    refinar este mecanismo para manejar de forma más elegante los errores de conversión.

    Args:
        df (pd.DataFrame): DataFrame con los datos.
        nombre_columna (str): Nombre de la columna a convertir.

    Returns:
        pd.DataFrame: DataFrame con la columna convertida a datetime.
    """
    def convertir_fecha(valor):
        """
        Intenta convertir un valor a datetime utilizando una lista de formatos predefinidos.
        Si el valor es nulo o indica ausencia de dato, devuelve np.nan.
        
        Args:
            valor: Valor a convertir.

        Returns:
            datetime o np.nan: Valor convertido o np.nan si no se pudo convertir.
        """
        if pd.isna(valor) or valor == '' or valor == '-' or valor == 'no se registro ?':
            return np.nan
        
        # Convertir el valor a string y eliminar espacios en blanco
        valor = str(valor).strip()
        
        # Lista de formatos de fecha a intentar
        formatos_fecha = [
            '%d/%m/%y %I:%M %p',  # Ejemplo: 03/02/25 10:44 AM
            '%Y-%m-%d %H:%M:%S',  # Ejemplo: 2025-02-03 09:45:00
            '%d/%m/%y %H:%M:%S',  # Ejemplo: 03/02/25 15:02:15
            '%d/%m/%y',           # Ejemplo: 03/02/25
            '%Y-%m-%d %H:%M',     # Ejemplo: 2025-02-03 09:48
            '%d/%m/%Y %I:%M %p',  # Ejemplo: 03/02/2025 07:38 AM
            '%d/%m/%Y %H:%M',     # Ejemplo: 31/01/2025 13:04
        ]
        
        # Intentar convertir el valor con cada formato
        for formato in formatos_fecha:
            try:
                return datetime.strptime(valor, formato)
            except ValueError:
                continue
        
        # Si ninguno de los formatos funciona, devuelve np.nan
        return np.nan

    valores_no_convertidos = []  # Lista para almacenar valores que no se pudieron convertir

    # Aplicar la función de conversión a la columna especificada.
    # Se utiliza "or valores_no_convertidos.append(x)" para capturar los valores no convertidos.
    columna_unificada = df[nombre_columna].apply(lambda x: convertir_fecha(x) or valores_no_convertidos.append(x))

    # Imprimir los primeros 20 valores que no se pudieron convertir, si los hubiera.
    if valores_no_convertidos:
        print(f"Columna: {nombre_columna} - Valores no convertidos: {', '.join(map(str, valores_no_convertidos[:20]))}")

    # Reemplazar la columna original con la columna convertida
    df[nombre_columna] = columna_unificada

    return df


def actualizar_tipos_datos(conexion, tabla, df):
    """
    Actualiza el tipo de dato de cada columna en el DataFrame según la información en la tabla de metadatos.

    La función realiza los siguientes pasos:
      - Lee la tabla de metadatos 'metadatos_de_tablas' para la tabla específica.
      - Verifica que todas las columnas del DataFrame existan en los metadatos.
      - Si faltan columnas en los metadatos, solicita al usuario que ingrese el tipo de dato adecuado.
      - Convierte las columnas del DataFrame al tipo de dato indicado:
            INTEGER, REAL, TEXT, DATE o DATETIME.

    Args:
        conexion (sqlite3.Connection): Conexión activa a la base de datos.
        tabla (str): Nombre de la tabla a actualizar.
        df (pd.DataFrame): DataFrame con los datos a actualizar.
    """
    tabla_de_columnas = 'metadatos_de_tablas'

    # Leer los metadatos para la tabla indicada
    query = f"SELECT * FROM {tabla_de_columnas} WHERE nombre_tabla = '{tabla}'"
    metadatos = pd.read_sql_query(query, conexion)

    # Verificar si hay columnas en el DataFrame que no están en los metadatos
    columnas_faltantes = [col for col in df.columns if col not in metadatos['nombre_columna'].values]

    # Para cada columna faltante, se solicita al usuario que ingrese el tipo de dato
    for columna in columnas_faltantes:
        while True:
            tipo_dato = input(f"Por favor, ingrese el tipo de dato para la columna '{columna}' \n(1: INTEGER, 2: REAL, 3: TEXT, 4: DATE, 5: DATETIME): ").strip()
            tipo_dato_map = {'1': 'INTEGER', '2': 'REAL', '3': 'TEXT', '4': 'DATE', '5': 'DATETIME'}
            if tipo_dato in tipo_dato_map:
                tipo_dato = tipo_dato_map[tipo_dato]
                print(f"Se ha seleccionado el tipo de dato '{tipo_dato}' para la columna '{columna}'.\n")
                nuevo_metadato = pd.DataFrame({'nombre_tabla': [tabla], 'nombre_columna': [columna], 'tipo_dato': [tipo_dato]})
                nuevo_metadato.to_sql(tabla_de_columnas, conexion, if_exists='append', index=False)
                break
            else:
                print("Tipo de dato no válido. Intente nuevamente.")
                
    # Actualizar la tabla de metadatos luego de agregar las nuevas columnas
    metadatos = pd.read_sql_query(query, conexion)

    # Convertir cada columna del DataFrame al tipo de dato correspondiente
    for _, row in metadatos.iterrows():
        nombre_columna = row['nombre_columna']
        tipo_dato = row['tipo_dato']
        
        if nombre_columna in df.columns:
            if tipo_dato == 'INTEGER':
                df[nombre_columna] = pd.to_numeric(df[nombre_columna], errors='coerce')
                # Si el resultado es float64, se redondea y se convierte a entero
                if df[nombre_columna].dtype == 'float64':
                    df[nombre_columna] = df[nombre_columna].round(0)
                df[nombre_columna] = df[nombre_columna].astype('Int64')
            elif tipo_dato == 'REAL':
                df[nombre_columna] = pd.to_numeric(df[nombre_columna], errors='coerce').astype(float)
            elif tipo_dato == 'TEXT':
                df[nombre_columna] = df[nombre_columna].astype(str)
            elif tipo_dato == 'DATE':
                df[nombre_columna] = pd.to_datetime(df[nombre_columna], errors='coerce').dt.date
            elif tipo_dato == 'DATETIME':
                df = convertir_fechas(df, nombre_columna)


def actualizar_base_datos(conexion, tabla, df, id):
    """
    Guarda el DataFrame en la tabla especificada de la base de datos SQLite.

    La función realiza lo siguiente:
      - Intenta leer la tabla existente para obtener datos previos.
      - Si la tabla existe, concatena los datos nuevos con los existentes y elimina duplicados según la columna 'id'.
      - Si la tabla no existe, procede a crearla.
      - Actualiza los tipos de datos de cada columna mediante la función 'actualizar_tipos_datos'.
      - Finalmente, guarda el DataFrame en la base de datos reemplazando la tabla existente.

    Args:
        conexion (sqlite3.Connection): Conexión activa a la base de datos.
        tabla (str): Nombre de la tabla donde se guardarán los datos.
        df (pd.DataFrame): DataFrame con los datos a insertar.
        id (str): Nombre de la columna utilizada para identificar duplicados. 
                  Si es 'Index', se crea una columna compuesta a partir de 'Order_ID' y 'Operation_Time'.
    """
    try:
        # Se intenta leer la tabla existente
        query = f"SELECT * FROM {tabla}"
        df_viejo = pd.read_sql_query(query, conexion)
        print(f"\tLa tabla {tabla} se actualiza.")
        # Se concatenan los datos existentes con los nuevos
        df = pd.concat([df_viejo, df], ignore_index=True)
        # Si la tabla incluye "TOA", limpiar la columna 'Estado_TOA'
        if "TOA" in tabla:
            df['Estado_TOA'] = df['Estado_TOA'].str.replace("(antiguo)", "", regex=False).str.strip()
            print("\tSe ha limpiado la columna 'Estado_TOA' de la tabla.")
        # Se eliminan duplicados basándose en la columna 'id'
        if id != 'Index':
            df = df.drop_duplicates(subset=id, keep='last')
        else:
            df['Index'] = df['Order_ID'].astype(str) + '_' + df['Operation_Time'].astype(str)
            df = df.drop_duplicates(subset=id, keep='last')
    except pd.io.sql.DatabaseError:
        # Si la tabla no existe, se utilizarán los datos nuevos para crearla
        print(f"\tLa tabla {tabla} no existe. Se creará una nueva.")

    # Actualizar los tipos de datos de cada columna según los metadatos
    actualizar_tipos_datos(conexion, tabla, df)

    # Guardar (o reemplazar) la tabla en la base de datos
    print(f"\tGuardando datos en la tabla {tabla}...")
    df.to_sql(tabla, conexion, if_exists='replace', index=False)
    print(f"\tTabla {tabla} actualizada correctamente.")


def marcar_archivos_procesados(carpeta, archivos):
    """
    Renombra los archivos Excel agregando el sufijo '_procesado' junto con la fecha actual
    y los mueve a una subcarpeta 'old' para indicar que ya fueron procesados.

    Args:
        carpeta (str): Ruta de la carpeta que contiene los archivos.
        archivos (list): Lista de nombres de archivos procesados.
    """
    # Se define la ruta de la carpeta 'old' donde se moverán los archivos procesados
    carpeta_old = os.path.join(carpeta, 'old')
    # Se crea la carpeta 'old' si no existe
    if not os.path.exists(carpeta_old):
        os.makedirs(carpeta_old)
    
    # Para cada archivo en la lista, se renombra y se mueve a la carpeta 'old'
    for archivo in archivos:
        ruta_antigua = os.path.join(carpeta, archivo)
        nombre_base, extension = os.path.splitext(archivo)
        # Se obtiene la fecha actual en formato YYYYMMDD
        dia_actual = datetime.now().strftime("%Y%m%d")
        nuevo_nombre = f"{nombre_base}_procesado_{dia_actual}{extension}"
        ruta_nueva = os.path.join(carpeta_old, nuevo_nombre)
        
        try:
            os.rename(ruta_antigua, ruta_nueva)
        except OSError as e:
            print(f"Error al mover y renombrar el archivo {archivo}: {e}")


def combinar_datos_sitios(carpeta, tabla, conexion, id):
    """
    Combina la información de varios archivos Excel en un solo DataFrame y lo guarda en la 
    base de datos SQLite mediante la función 'actualizar_base_datos'.

    La función busca tres tipos de archivos dentro de la carpeta:
      - Archivos que contengan "sitios" en su nombre, que se consideran como base_sitios.
      - Archivos que contengan "swap" en su nombre, de los cuales se extraen únicamente las columnas
        ["Codigo Estacion", "Fecha Fin Swap", "Alarmas Activas Nodo"].
      - Archivos que contengan "tss" en su nombre, asignando a casos_tss y mostrando la cantidad de filas.

    Se renombrarán algunas columnas para unificar criterios y se fusionarán los DataFrames 
    resultantes en base a la columna 'Codigo Unico'.

    Args:
        carpeta (str): Ruta de la carpeta que contiene los archivos.
        tabla (str): Nombre de la tabla en la base de datos donde se guardarán los datos.
        conexion (sqlite3.Connection): Conexión activa a la base de datos SQLite.
        id (str): Columna identificadora para evitar duplicados al actualizar la base de datos.

    Raises:
        ValueError: Si no se encuentran todos los archivos necesarios (sitios, swap y tss).
    """
    base_sitios, casos_swap, casos_tss = None, None, None

    # Recorrer cada archivo en la carpeta para identificar y leer los archivos requeridos
    for file in os.listdir(carpeta):
        file_path = os.path.join(carpeta, file)
        
        if "sitios" in file.lower():
            base_sitios = pd.read_excel(file_path)
        elif "swap" in file.lower():
            casos_swap = pd.read_excel(file_path)
            # Se filtran únicamente las columnas relevantes
            casos_swap = casos_swap[["Codigo Estacion", "Fecha Fin Swap", "Alarmas Activas Nodo"]]
        elif "tss" in file.lower():
            casos_tss = pd.read_excel(file_path)

    # Validar que se hayan encontrado los tres archivos requeridos
    if base_sitios is None or casos_swap is None or casos_tss is None:
        raise ValueError("No se encontraron todos los archivos necesarios en la carpeta.")

    # Renombrar columnas para unificar la llave de unión
    casos_swap.rename(columns={"Codigo Estacion": "Codigo Unico"}, inplace=True)
    casos_tss.rename(columns={"Customer Site ID": "Codigo Unico"}, inplace=True)
    # Se elimina el sufijo '_Swap' en los valores de la columna 'Codigo Unico' en casos_tss
    casos_tss["Codigo Unico"] = casos_tss["Codigo Unico"].str.replace("_Swap", "")

    # Fusionar base_sitios con casos_swap utilizando un join de tipo left
    df_merged = base_sitios.merge(casos_swap, on="Codigo Unico", how="left")
    # Fusionar el resultado anterior con casos_tss
    df_merged = df_merged.merge(casos_tss, on="Codigo Unico", how="left")
    # Reemplazar espacios en los nombres de columna por guiones bajos para estandarizar
    df_merged.columns = df_merged.columns.str.replace(' ', '_')

    # Actualizar la base de datos con el DataFrame resultante
    actualizar_base_datos(conexion, tabla, df_merged, id)


def actualizar_lista_tickets_test(conexion, tabla_tickets, tabla_tickets_test):
    """
    Actualiza la lista de tickets TEST en la base de datos y en el archivo Excel 'Tickets_cambios.xlsx'.

    La función realiza los siguientes pasos:
      - Obtiene tickets de la tabla principal que contengan términos de prueba en la columna 'Notas'.
      - Lee la hoja 'TEST' y 'ERRORES' del archivo Excel 'Tickets_cambios.xlsx'.
      - De la hoja 'ERRORES' filtra los registros que tienen 'TEST' en la columna 'Sustituido', 
        asignándoles valores por defecto para 'Notas' y 'Confirmado'.
      - Combina los DataFrames resultantes y elimina duplicados basados en 'Nro_TOA'.
      - Para los tickets que no tienen confirmación, solicita al usuario confirmar cada uno.
      - Guarda el DataFrame combinado en la base de datos y actualiza la hoja 'TEST' del archivo Excel.

    Args:
        conexion (sqlite3.Connection): Conexión activa a la base de datos.
        tabla_tickets (str): Nombre de la tabla de tickets en la base de datos.
        tabla_tickets_test (str): Nombre de la tabla de tickets TEST en la base de datos.
    """
    # Obtener tickets de la tabla principal que contengan "test" o "ticket de prueba" en 'Notas'
    query_tickets = f"SELECT Nro_TOA, Notas FROM {tabla_tickets} WHERE LOWER(Notas) LIKE '%test%' OR LOWER(Notas) LIKE '%ticket de prueba%'"
    df_tickets = pd.read_sql_query(query_tickets, conexion)

    # Leer la hoja 'TEST' del archivo Excel 'Tickets_cambios.xlsx'
    ruta_tickets_cambios = os.path.join(base_path, "DATA", 'INFO TICKETS', 'Tickets_cambios.xlsx')
    df_tickets_test = pd.read_excel(ruta_tickets_cambios, sheet_name='TEST')

    # Leer la hoja 'ERRORES' del archivo Excel y filtrar los registros que corresponden a TEST
    df_errores = pd.read_excel(ruta_tickets_cambios, sheet_name='ERRORES')
    df_errores_test = df_errores[df_errores['Sustituido'] == 'TEST']
    df_errores_test = df_errores_test[['Nro_TOA']].copy()
    df_errores_test['Notas'] = 'Error al pasar de Autin'
    df_errores_test['Confirmado'] = 'SI'

    # Combinar los DataFrames y eliminar duplicados basados en 'Nro_TOA'
    df_combined = pd.concat([df_tickets, df_tickets_test, df_errores_test], ignore_index=True)
    df_combined = df_combined.drop_duplicates(subset='Nro_TOA', keep='last')

    # Filtrar tickets que aún necesitan confirmación (donde 'Confirmado' es NaN)
    df_necesitan_confirmacion = df_combined[df_combined['Confirmado'].isna()]

    # Mostrar la cantidad de tickets que necesitan confirmación
    tickets_por_confirmar = len(df_necesitan_confirmacion)
    if tickets_por_confirmar > 0:
        print(f"Hay {tickets_por_confirmar} tickets que necesitan confirmación.\n")

    # Para cada ticket pendiente, solicitar confirmación al usuario
    for index, row in df_necesitan_confirmacion.iterrows():
        nro_toa = row['Nro_TOA']
        notas = row['Notas']
        confirmacion = input(f"¿Confirma el ticket Nro TOA {nro_toa} con las notas '{notas}'? (S/N): ").strip().upper()
        tickets_por_confirmar -= 1
        print(f"Quedan {tickets_por_confirmar} tickets por confirmar.")
        if confirmacion == 'S':
            df_combined.at[index, 'Confirmado'] = 'SI'
        else:
            df_combined.at[index, 'Confirmado'] = 'NO'

    # Guardar los datos combinados en la tabla de tickets_test en la base de datos
    df_combined.to_sql(tabla_tickets_test, conexion, if_exists='replace', index=False)
    print(f"\tTabla {tabla_tickets_test} actualizada correctamente.")
    
    # Actualizar la hoja 'TEST' en el archivo Excel con el DataFrame combinado
    with pd.ExcelWriter(ruta_tickets_cambios, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_combined.to_excel(writer, sheet_name='TEST', index=False)


def actulizar_columnas(df):
    """
    Renombra y filtra las columnas de un DataFrame según un diccionario predefinido de mapeo.

    Se utiliza para estandarizar los nombres de las columnas a lo largo del proyecto,
    eliminando columnas que no se encuentren en el diccionario.

    Args:
        df (pd.DataFrame): DataFrame original con las columnas a renombrar.

    Returns:
        pd.DataFrame: DataFrame filtrado y con los nombres de columnas actualizados.
    """
    # Diccionario que mapea los nombres originales a los nuevos nombres estandarizados
    columnas_renombradas = {
        "Nro_TOA": "ID_TOA",
        "Fecha_de_Registro_de_actividad_TOA": "Creacion_TOA",
        "Fecha_Hora_de_Cancelación": "Cierre_TOA",
        "Subtipo_de_Actividad": "Tipo_Actividad",
        "Número_de_Petición": "Numero_Peticion",
        "ID_del_Ticket": "ID_Ticket",
        "SLA_Inicio": "SLA_Inicio",
        "SLA_Fin": "SLA_Fin",
        "Direccion_Polar_X": "Coordenada_X",
        "Direccion_Polar_Y": "Coordenada_Y",
        "Notas": "Notas",
        "Codigo_Unico": "Site_ID",
        "Nombre_Local_sitios": "Nombre_Local",
        "Empresa": "Empresa",
        "Bucket_Inicial": "Bucket",
        "Inicio_PR1": "Inicio_PR1",
        "Fin_PR1": "Fin_PR1",
        "Motivo_PR1": "Motivo_PR1",
        "Inicio_PR2": "Inicio_PR2",
        "Fin_PR2": "Fin_PR2",
        "Motivo_PR2": "Motivo_PR2",
        "Inicio_PR3": "Inicio_PR3",
        "Fin_PR3": "Fin_PR3",
        "Motivo_PR3": "Motivo_PR3",
        "Inicio_PR4": "Inicio_PR4",
        "Fin_PR4": "Fin_PR4",
        "Motivo_PR4": "Motivo_PR4",
        "Departamento": "Departamento",
        "Provincia": "Provincia",
        "Distrito": "Distrito",
        "Tipo_Local": "Tipo_Local",
        "Atencion": "Tipo_Atencion",
        "Zona_sitios": "Zona",
        "Tipo_Zona_FLM": "Tipo_Zona",
        "Tipo_Estacion": "Tipo_Estacion",
        "SLA": "SLA",
        "ubigeotoa": "Ubigeo_TOA",
        "Estado_TOA": "Estado_TOA",
        "Proactivo": "Proactivo",
        "Marcha_Blanca": "Marcha_Blanca",
        "Responsable": "Responsable",
        "TEST": "Test",
        "Fecha_Fin_Swap": "Fecha_Fin_Swap",
        "Alarmas_Activas_Nodo": "Alarmas_Activas",
        "SWAP_dias": "Dias_Swap",
        "Fecha_TSS": "Fecha_TSS",
        "TSS_dias": "Dias_TSS",

        "Task_Id_1": "Autin_ID_1",
        "Task_Status_1": "Estado_1",
        "Cancel_Reason_1": "Motivo_Cancel_1",
        "Hora_PR_1": "Hora_PR_1",
        "Motivo_PR_1": "Motivo_PR_1",
        "Estado_PR_1": "Estado_PR_1",

        "Task_Id_2": "Autin_ID_2",
        "Task_Status_2": "Estado_2",
        "Cancel_Reason_2": "Motivo_Cancel_2",
        "Hora_PR_2": "Hora_PR_2",
        "Motivo_PR_2": "Motivo_PR_2",
        "Estado_PR_2": "Estado_PR_2",

        "Task_Id_3": "Autin_ID_3",
        "Task_Status_3": "Estado_3",
        "Cancel_Reason_3": "Motivo_Cancel_3",
        "Hora_PR_3": "Hora_PR_3",
        "Motivo_PR_3": "Motivo_PR_3",
        "Estado_PR_3": "Estado_PR_3",

        "Tiempo_TOA_Autin": "Tiempo_TOA_Autin",
        
        "Task_Id_Abastecimiento_1": "Tarea_Abastecimiento",
        "Task_Status_Abastecimiento_1": "Estado_Abastecimiento",
        "Createtime_Abastecimiento_1": "Hora_Creacion_Abastecimiento",
        "Abastecimiento_dias_1": "Dias_Abastecimiento",

        "Reject_Counter_1": "Rechazos",        
        "Com_Level_1_Aff_Equip_1": "Equipo_Afectado",
        "Duration_hours_1": "Duracion_Horas",

        "Reiteradas": "Reiteradas",
        "TOA_Reiterado": "TOA_Reiterdo"
    }

    # Filtrar únicamente las columnas que están definidas en el diccionario
    df_filtrado = df[list(columnas_renombradas.keys())]
    # Renombrar las columnas usando el mapeo definido
    df_filtrado = df_filtrado.rename(columns=columnas_renombradas)

    return df_filtrado


def etiquetar_nro_toa_y_rango(df_merged, archivo_excel):
    """
    Etiqueta los registros del DataFrame 'df_merged' de la siguiente forma:
      - Crea la columna 'EN_TDE' asignando 'SI' si el 'ID_TOA' se encuentra en la columna
        'activityId' del archivo Excel; de lo contrario, asigna 'NO'.
      - Etiqueta cada registro con la etiqueta 'en_rango' en la columna 'en_rango' si la fecha 
        en 'Creacion_TOA' se encuentra dentro del rango definido por la fecha mínima y máxima de 
        la columna 'timeOfBooking' del archivo Excel; de lo contrario, deja una cadena vacía.

    Args:
        df_merged (pd.DataFrame): DataFrame que contiene la información consolidada.
        archivo_excel (str): Ruta del archivo Excel que contiene la información de referencia 
                             (debe incluir las columnas 'activityId' y 'timeOfBooking').

    Returns:
        pd.DataFrame: DataFrame con las nuevas columnas 'EN_TDE' y 'en_rango' actualizadas.
    """
    # Leer el archivo Excel que contiene los datos de referencia
    df_excel = pd.read_excel(archivo_excel)

    # Asegurar que las columnas 'ID_TOA' y 'activityId' sean de tipo string para la comparación
    df_merged['ID_TOA'] = df_merged['ID_TOA'].astype(str)
    df_excel['activityId'] = df_excel['activityId'].astype(str)

    # Crear la columna 'EN_TDE': 'SI' si el ID_TOA se encuentra en activityId, de lo contrario 'NO'
    df_merged['EN_TDE'] = df_merged['ID_TOA'].isin(df_excel['activityId']).map({True: 'SI', False: 'NO'})

    # Convertir la columna 'timeOfBooking' a datetime, considerando el formato esperado
    df_excel['timeOfBooking'] = pd.to_datetime(df_excel['timeOfBooking'], format='%Y-%m-%d %H:%M:%S', errors='coerce')
    
    # Definir el rango de fechas utilizando la fecha mínima y máxima de 'timeOfBooking'
    fecha_max = df_excel['timeOfBooking'].max()
    fecha_min = df_excel['timeOfBooking'].min()

    # Etiquetar cada registro de df_merged con 'en_rango' si 'Creacion_TOA' se encuentra dentro del rango
    df_merged['en_rango'] = df_merged['Creacion_TOA'].apply(lambda x: 'en_rango' if fecha_min <= x <= fecha_max else '')

    return df_merged


def combinar_tablas(conexion, tabla_TOA, tabla_autin, tabla_sitios, tabla_final):
    """
    Combina las tablas TOA, autin y sitios de la base de datos en una tabla consolidada.
    
    La función realiza los siguientes pasos:
      1. Lee las tablas TOA, autin y sitios desde la base de datos.
      2. Actualiza los tipos de datos en cada DataFrame usando metadatos.
      3. Renombra la columna 'Código_de_Cliente' a 'Codigo_Unico' en TOA para facilitar la unión.
      4. Verifica columnas duplicadas entre TOA y sitios (excepto 'Codigo_Unico') y las renombra en sitios.
      5. Une TOA y sitios usando 'Codigo_Unico' y ordena por 'Fecha_de_Registro_de_actividad_TOA'.
      6. Actualiza la lista de tickets test y marca en TOA los tickets confirmados como test.
      7. Asigna valores en la columna 'Empresa' según si 'Bucket_Inicial' contiene "comfica" o "huawei".
      8. Marca 'Marcha_Blanca' si el Departamento es "Puno" o la Provincia es "Cañete".
      9. Marca como "Proactivo" en función de si en las Notas aparece la palabra "proactivo".
      10. Define el responsable como 'FLM' si el Bucket contiene "comfica" o "huawei"; de lo contrario, 'TDP'.
      11. Convierte columnas de fechas y calcula los días de diferencia para SWAP y TSS.
      12. Clasifica los tickets de Autin, renombra la columna de unión y la une con TOA.
      13. Calcula el tiempo entre la creación de TOA y el ticket de Autin (en minutos).
      14. Identifica tickets reiterados (tickets repetidos en un lapso de 7 días) y asigna la referencia del ticket anterior.
      15. Estandariza los nombres de las columnas y etiqueta los tickets en función de un rango definido en un archivo Excel.
      16. Asigna etiquetas personalizadas según condiciones en 'Estado_TOA' y 'Estado_1'.
      17. Actualiza la tabla final en la base de datos y genera archivos Excel separados para Comfica y Huawei.
    
    Args:
        conexion (sqlite3.Connection): Conexión activa a la base de datos.
        tabla_TOA (str): Nombre de la tabla TOA en la base de datos.
        tabla_autin (str): Nombre de la tabla autin en la base de datos.
        tabla_sitios (str): Nombre de la tabla sitios en la base de datos.
        tabla_final (str): Nombre de la tabla consolidada a crear/actualizar.
    """
    # 1. Leer las tablas desde la base de datos
    query_TOA = f"SELECT * FROM {tabla_TOA}"
    query_autin = f"SELECT * FROM {tabla_autin}"
    query_sitios = f"SELECT * FROM {tabla_sitios}"

    df_TOA = pd.read_sql_query(query_TOA, conexion)
    df_autin = pd.read_sql_query(query_autin, conexion)
    df_sitios = pd.read_sql_query(query_sitios, conexion)

    # 2. Actualizar los tipos de datos en cada DataFrame según los metadatos
    actualizar_tipos_datos(conexion, tabla_TOA, df_TOA)
    actualizar_tipos_datos(conexion, tabla_autin, df_autin)
    actualizar_tipos_datos(conexion, tabla_sitios, df_sitios)

    # 3. Renombrar 'Código_de_Cliente' a 'Codigo_Unico' en TOA para facilitar la unión
    df_TOA.rename(columns={'Código_de_Cliente': 'Codigo_Unico'}, inplace=True)

    # 4. Verificar columnas duplicadas entre TOA y sitios (excepto 'Codigo_Unico') y renombrarlas en sitios
    columnas_repetidas = [col for col in df_TOA.columns if col in df_sitios.columns and col != 'Codigo_Unico']
    if columnas_repetidas:
        for col in columnas_repetidas:
            df_sitios.rename(columns={col: f"{col}_sitios"}, inplace=True)

    # 5. Unir df_TOA y df_sitios usando 'Codigo_Unico' y ordenar por 'Fecha_de_Registro_de_actividad_TOA'
    df_merged = df_TOA.merge(df_sitios, on="Codigo_Unico", how="left")    
    df_merged.sort_values(by='Fecha_de_Registro_de_actividad_TOA', inplace=True)

    # 6. Actualizar la lista de tickets test y marcar en TOA los tickets confirmados como test
    actualizar_lista_tickets_test(conexion, tabla_TOA, 'tickets_test')
    df_tickets_test = pd.read_sql_query("SELECT Nro_TOA FROM tickets_test WHERE Confirmado = 'SI'", conexion)
    nro_toa_tickets_test = df_tickets_test['Nro_TOA'].tolist()
    df_merged['TEST'] = df_merged['Nro_TOA'].apply(lambda x: 'TEST' if x in nro_toa_tickets_test else '')

    # 7. Asignar la columna 'Empresa' según el contenido de 'Bucket_Inicial'
    df_merged['Empresa'] = df_merged['Bucket_Inicial'].apply(
        lambda x: 'comfica' if 'comfica' in x.lower() else ('huawei' if 'huawei' in x.lower() else '')
    )
    
    # 8. Marcar como 'MB' en 'Marcha_Blanca' si Departamento es "Puno" o Provincia es "Cañete"
    df_merged['Marcha_Blanca'] = df_merged.apply(
        lambda x: 'MB' if x['Departamento'] == 'Puno' or x['Provincia'] == 'Cañete' else '', axis=1
    )

    # 9. Marcar como "Proactivo" si en 'Notas' aparece la palabra "proactivo"
    df_merged['Notas'] = df_merged['Notas'].fillna('')
    df_merged['Proactivo'] = df_merged['Notas'].str.contains('proactivo', case=False, na=False).map({True: 'Proactivo', False: ''})

    # 10. Asignar 'Responsable': 'FLM' si 'Bucket_Inicial' contiene "comfica" o "huawei", de lo contrario 'TDP'
    df_merged['Responsable'] = df_merged['Bucket_Inicial'].apply(
        lambda x: 'FLM' if 'comfica' in x or 'huawei' in x else 'TDP'
    )

    # 11. Convertir las columnas de fechas para SWAP y TSS a datetime y calcular la diferencia en días
    df_merged['Fecha_Fin_Swap'] = pd.to_datetime(df_merged['Fecha_Fin_Swap'], format='%Y-%m-%d %H:%M:%S', errors='coerce')
    df_merged['Fecha_TSS'] = pd.to_datetime(df_merged['Fecha_TSS'], errors='coerce')
    df_merged['SWAP_dias'] = df_merged.apply(
        lambda x: (x['Fecha_de_Registro_de_actividad_TOA'] - x['Fecha_Fin_Swap']).days if (x['Fecha_de_Registro_de_actividad_TOA'] - x['Fecha_Fin_Swap']).days > 0 else None,
        axis=1
    )
    df_merged['TSS_dias'] = df_merged.apply(
        lambda x: (x['Fecha_de_Registro_de_actividad_TOA'] - x['Fecha_TSS']).days if (x['Fecha_de_Registro_de_actividad_TOA'] - x['Fecha_TSS']).days > 0 else None,
        axis=1
    )

    # 12. Clasificar los tickets de Autin y renombrar 'Number_OS_SIOM' a 'Nro_TOA' para la unión
    df_autin = clasificar_tickets_autin(df_autin, conexion)
    df_autin.rename(columns={'Number_OS_SIOM': 'Nro_TOA'}, inplace=True)

    # 13. Asegurar formato string y sin espacios en 'Nro_TOA' en ambos DataFrames, y unirlos
    df_merged['Nro_TOA'] = df_merged['Nro_TOA'].astype(str).str.strip()
    df_autin['Nro_TOA'] = df_autin['Nro_TOA'].astype(str).str.strip()
    df_merged = df_merged.merge(df_autin, on='Nro_TOA', how='left')
    print(f"\tSe han combinado los DataFrames de TOA y Autin. Tamaño: {df_merged.shape}")

    # 14. Calcular 'Tiempo_TOA_Autin' en minutos
    df_merged['Tiempo_TOA_Autin'] = (df_merged['Createtime_1'] - df_merged['Fecha_de_Registro_de_actividad_TOA']).dt.total_seconds() / 60

    # 15. Identificar tickets reiterados en un lapso de 7 días y asignar referencia
    df_merged.sort_values(by=['Codigo_Unico', 'Com_Level_1_Aff_Equip_1', 'Fecha_de_Registro_de_actividad_TOA'], inplace=True)
    df_merged['Reiteradas'] = ''
    df_merged['TOA_Reiterado'] = ''
    for (codigo_unico, equip), group in df_merged.groupby(['Codigo_Unico', 'Com_Level_1_Aff_Equip_1']):
        if equip:  # Se evalúa solo si 'Com_Level_1_Aff_Equip_1' tiene valor
            for i in range(1, len(group)):
                current_ticket = group.iloc[i]
                previous_tickets = group.iloc[:i]
                previous_ticket = previous_tickets[previous_tickets['Fecha_de_Registro_de_actividad_TOA'] >= current_ticket['Fecha_de_Registro_de_actividad_TOA'] - pd.Timedelta(days=7)]
                if not previous_ticket.empty:
                    df_merged.loc[current_ticket.name, 'Reiteradas'] = 'Reiterada'
                    df_merged.loc[current_ticket.name, 'TOA_Reiterado'] = previous_ticket.iloc[-1]['Nro_TOA']

    # 16. Estandarizar nombres de columnas mediante la función 'actulizar_columnas'
    df_merged = actulizar_columnas(df_merged)
    
    # 17. Ordenar el DataFrame por 'Creacion_TOA' en orden descendente
    df_merged.sort_values(by='Creacion_TOA', ascending=False, inplace=True)

    # 18. Etiquetar tickets en función de un rango definido en un archivo Excel
    archivo_excel = os.path.join(base_path, "REPORTES TDE", "PINT_Reporte_Mtto_Correctivo.xlsx")
    df_merged = etiquetar_nro_toa_y_rango(df_merged, archivo_excel)

    # 19. Asignar etiquetas personalizadas según condiciones en 'Estado_TOA' y 'Estado_1'
    def asignar_etiqueta(row):
        if row['Estado_TOA'] in ['Cancelado'] and row['Estado_1'] in ['accepted', 'closed', 'completed', 'dispatched', 'inprocess', 'unscheduled']:
            return 'cruce incorrecto'
        elif row['Estado_TOA'] in ['Completado'] and row['Estado_1'] in ['accepted', 'canceled', 'dispatched', 'inprocess', 'unscheduled']:
            return 'cruce incorrecto'
        elif row['Estado_TOA'] in ['Pendiente'] and row['Estado_1'] in ['closed', 'completed', 'canceled']:
            return 'falta de revisión'
        elif row['Estado_TOA'] in ['Pre cierre'] and row['Estado_1'] in ['closed', 'completed', 'canceled']:
            return 'falta de revisión'
        elif row['Estado_TOA'] in ['Suspendido'] and row['Estado_1'] in ['closed', 'completed', 'canceled']:
            return 'falta de revisión'
        else:
            return ''
    df_merged['Etiqueta'] = df_merged.apply(asignar_etiqueta, axis=1)

    # 20. Actualizar la tabla final consolidada en la base de datos
    actualizar_base_datos(conexion, tabla_final, df_merged, 'ID_TOA')

    # 21. Generar archivos Excel separados para Comfica y Huawei
    columnas_a_eliminar = [
        'Tarea_Abastecimiento', 'Estado_Abastecimiento', 'Hora_Creacion_Abastecimiento', 
        'Dias_Abastecimiento', 'Rechazos', 'Equipo_Afectado', 'Duracion_Horas', 
        'Reiteradas', 'TOA_Reiterdo', 'EN_TDE', 'en_rango', 'Proactivo', 'Marcha_Blanca', 
        'Responsable', 'Test', 'Fecha_Fin_Swap', 'Alarmas_Activas', 'Dias_Swap', 
        'Fecha_TSS', 'Dias_TSS', 'Etiqueta'
    ]
    df_comfica = df_merged[df_merged['Bucket'].str.contains('comfica', case=False, na=False)].copy()
    archivo_comfica = os.path.join(base_path, "reporte_Comfica.xlsx")
    df_comfica.replace([0, np.nan, None, 'nan'], '', inplace=True)
    df_comfica.drop(columns=columnas_a_eliminar, inplace=True, errors='ignore')
    df_comfica.to_excel(archivo_comfica, index=False)
    print(f"\tArchivo Excel para Comfica guardado en: {archivo_comfica}")

    df_huawei = df_merged[df_merged['Bucket'].str.contains('huawei', case=False, na=False)].copy()
    archivo_huawei = os.path.join(base_path, "reporte_Huawei.xlsx")
    df_huawei.replace([0, np.nan, None, 'nan'], '', inplace=True)
    df_huawei.drop(columns=columnas_a_eliminar, inplace=True, errors='ignore')
    df_huawei.to_excel(archivo_huawei, index=False)
    print(f"\tArchivo Excel para Huawei guardado en: {archivo_huawei}")


def ordenar_y_seleccionar_tickets(grupo, max_tickets):
    """
    Ordena los tickets de un grupo según una clave de prioridad y la fecha de creación, y selecciona
    los primeros 'max_tickets' tickets.

    La prioridad se define en base al estado de la tarea:
      - 'closed' tiene la mayor prioridad (1)
      - 'completed' es la segunda prioridad (2)
      - 'canceled' es la menor prioridad (4)
      - Estados no definidos reciben prioridad 3

    Args:
        grupo (pd.DataFrame): Grupo de tickets a ordenar.
        max_tickets (int): Número máximo de tickets a seleccionar.

    Returns:
        pd.DataFrame: DataFrame con hasta 'max_tickets' tickets ordenados.
    """
    if grupo.empty:
        return pd.DataFrame()

    # Asignar prioridad basada en 'Task_Status'
    prioridad = {
        'closed': 1,
        'completed': 2,
        'canceled': 4  # 'canceled' es de menor prioridad
    }
    grupo['Prioridad'] = grupo['Task_Status'].map(prioridad).fillna(3)
    # Ordenar por prioridad y por 'Createtime'
    grupo = grupo.sort_values(by=['Prioridad', 'Createtime'], ascending=[True, True])
    grupo['Orden'] = range(1, len(grupo) + 1)
    # Seleccionar los primeros max_tickets
    return grupo.head(max_tickets)


def clasificar_tickets_autin(df_autin, conexion):
    """
    Clasifica los tickets de Autin y les asigna una prioridad, consolidando información adicional
    de tickets de abastecimiento y de PR.

    Pasos realizados:
      1. Leer la tabla 'tickets_pr' desde la base de datos y renombrar sus columnas.
      2. Convertir la columna 'Createtime' a datetime y filtrar las columnas relevantes de df_autin.
      3. Verificar duplicados en 'Task_Id' y notificar si existen.
      4. Filtrar los tickets relacionados con "Abastecimiento" (excluyendo cancelados) y renombrar columnas.
      5. Excluir tickets con ciertas razones de cancelación y filtrar por patrones en 'Task_Id' y 'Task_Category'.
      6. Unir la información de abastecimiento con df_autin y calcular la diferencia en días.
      7. Unir la información de PR proveniente de 'tickets_pr'.
      8. Agrupar por 'Number_OS_SIOM', ordenar y seleccionar hasta 3 tickets por grupo.
      9. Calcular la duración en horas entre 'Complete_Time' y 'Createtime'.
      10. Pivotar el DataFrame para tener un ticket por fila y asegurar que existan las columnas esperadas.

    Args:
        df_autin (pd.DataFrame): DataFrame con los datos de los tickets de Autin.
        conexion (sqlite3.Connection): Conexión activa a la base de datos.

    Returns:
        pd.DataFrame: DataFrame final consolidado y pivotado con la clasificación de tickets de Autin.
    """
    # 1. Leer y renombrar columnas de la tabla 'tickets_pr'
    tabla_pr = 'tickets_pr'
    query = f"SELECT * FROM {tabla_pr}"
    df_tickets_pr = pd.read_sql_query(query, conexion)[['Order_ID', 'Operation_Time', 'Pause_Time', 'Reason']]
    df_tickets_pr.rename(columns={'Order_ID': 'Task_Id', 'Operation_Time': 'Hora_PR', 'Pause_Time': 'Estado_PR', 'Reason': 'Motivo_PR'}, inplace=True)

    # 2. Convertir 'Createtime' a datetime y seleccionar columnas relevantes de df_autin
    df_autin['Createtime'] = pd.to_datetime(df_autin['Createtime'], format='%Y-%m-%d %H:%M:%S', errors='coerce')
    df_autin = df_autin[['Task_Id', 'Task_Category', 'Createtime', 'Cancel_Reason', 'Task_Status', 'Site_Id', 'Com_Level_1_Aff_Equip', 'Number_OS_SIOM', 'Reject_Counter', 'Complete_Time']]
    
    # 3. Comprobar duplicados en 'Task_Id'
    if df_autin['Task_Id'].duplicated().any():
        print("Hay Task Id duplicados en el dataframe de Autin")

    # 4. Filtrar tickets de "Abastecimiento" (no cancelados) y renombrar columnas
    Autin_abastecimiento = df_autin[
        (df_autin['Task_Category'].str.contains("Abastecimiento", case=False, na=False)) &
        (df_autin['Task_Status'] != "canceled")
    ][['Site_Id', 'Task_Id', 'Task_Status', 'Createtime']]
    Autin_abastecimiento.columns = ['Site_Id', 'Task_Id_Abastecimiento', 'Task_Status_Abastecimiento', 'Createtime_Abastecimiento']
    Autin_abastecimiento.sort_values(by=['Site_Id', 'Createtime_Abastecimiento'], inplace=True)

    # 5. Excluir tickets con ciertas razones de cancelación y aplicar filtros basados en 'Task_Id' y 'Task_Category'
    df_autin = df_autin[(df_autin['Cancel_Reason'] != 'Duplicado') &
                        (df_autin['Cancel_Reason'] != 'Other') &
                        (df_autin['Cancel_Reason'] != 'Tarea de prueba') &
                        (df_autin['Cancel_Reason'] != 'Monitoreo')]
    df_autin = df_autin[
        (df_autin['Task_Id'].str.contains("CM", case=False, na=False)) |
        ((df_autin['Task_Id'].str.contains("PLM", case=False, na=False)) & (df_autin['Task_Category'] == "PROACTIVO"))
    ]

    # 6. Unir la información de abastecimiento con df_autin basándose en 'Site_Id'
    df_autin_con_abastecimiento = df_autin.merge(Autin_abastecimiento, on='Site_Id', how='left')
    df_autin_con_abastecimiento = df_autin_con_abastecimiento[df_autin_con_abastecimiento['Createtime_Abastecimiento'] < df_autin_con_abastecimiento['Createtime']]
    df_autin_con_abastecimiento.sort_values(by=['Site_Id', 'Createtime_Abastecimiento'], inplace=True)
    df_autin_con_abastecimiento.drop_duplicates(subset=['Site_Id', 'Task_Id'], keep='last', inplace=True)
    df_autin = pd.concat([df_autin, df_autin_con_abastecimiento], ignore_index=True)
    df_autin.drop_duplicates(subset='Task_Id', keep='last', inplace=True)
    df_autin['Abastecimiento_dias'] = (df_autin['Createtime'] - df_autin['Createtime_Abastecimiento']).dt.days

    # 7. Unir la información de PR proveniente de 'tickets_pr'
    df_autin = df_autin.merge(df_tickets_pr, on='Task_Id', how='left')

    # 8. Agrupar por 'Number_OS_SIOM', ordenar y seleccionar hasta 3 tickets por grupo
    df_autin['Number_OS_SIOM'] = df_autin['Number_OS_SIOM'].astype(str)
    df_autin = df_autin[df_autin['Number_OS_SIOM'].str.len() == 8]
    df_autin = df_autin.groupby('Number_OS_SIOM', group_keys=False).apply(lambda x: ordenar_y_seleccionar_tickets(x, 3))
    
    # 9. Convertir 'Complete_Time' a datetime y calcular la duración en horas
    df_autin['Complete_Time'] = pd.to_datetime(df_autin['Complete_Time'], format='%Y-%m-%d %H:%M:%S', errors='coerce')
    df_autin['Duration_hours'] = df_autin.apply(
        lambda row: (row['Complete_Time'] - row['Createtime']).total_seconds() / 3600 if pd.notnull(row['Complete_Time']) else None,
        axis=1
    )

    # 10. Reordenar y pivotar el DataFrame para tener un ticket por fila
    df_autin = df_autin.sort_values(by=['Number_OS_SIOM', 'Orden'])
    df_autin['Orden_Index'] = df_autin.groupby('Number_OS_SIOM').cumcount() + 1
    df_pivot = df_autin.pivot(index='Number_OS_SIOM', columns='Orden_Index')
    df_pivot.columns = [f"{col[0]}_{col[1]}" for col in df_pivot.columns]
    df_final = df_pivot.reset_index()

    # 11. Asegurar que las columnas esperadas existan, añadiéndolas si faltan
    columnas_esperadas = [
        'Task_Id_3', 'Task_Status_3', 'Cancel_Reason_3',
        'Hora_PR_3', 'Motivo_PR_3', 'Estado_PR_3'
    ]
    for col in columnas_esperadas:
        if col not in df_final.columns:
            df_final[col] = pd.NA 

    return df_final


def convertir_tabla_a_excel(tabla, archivo_salida, conexion, hoja_nombre='Sheet1'):
    """
    Convierte una tabla de la base de datos en un archivo Excel formateado.

    La función realiza lo siguiente:
      - Lee la tabla desde la base de datos y actualiza sus tipos de datos.
      - Crea un nuevo libro de trabajo y define una hoja con el nombre especificado.
      - Agrega los datos del DataFrame a la hoja, aplicando formatos condicionales:
            * Para celdas con valores numéricos (ej. "Rechazos", "Dias_Swap", "Dias_TSS") se aplica un formato naranja si cumplen ciertas condiciones.
            * Para celdas correspondientes a "Estado_PR_X", si contienen la palabra "Pause", también se formatean en naranja.
      - Agrega una columna con una fórmula de Excel que calcula el "Semaforo" (diferencia en horas entre la fecha actual y un valor en la columna B).
      - Define y crea una tabla de Excel con estilo y aplica formato condicional a la columna "Semaforo".
      - Guarda y cierra el archivo Excel.

    Args:
        tabla (str): Nombre de la tabla en la base de datos.
        archivo_salida (str): Ruta completa del archivo Excel de salida.
        conexion (sqlite3.Connection): Conexión activa a la base de datos SQLite.
        hoja_nombre (str, opcional): Nombre de la hoja en el archivo Excel. Por defecto "Sheet1".
    """
    # Crear un nuevo libro de trabajo y establecer la hoja activa con el nombre indicado
    wb = Workbook()
    ws = wb.active
    ws.title = hoja_nombre

    # Leer la tabla desde la base de datos y actualizar los tipos de datos usando los metadatos
    df = pd.read_sql_query(f"SELECT * FROM {tabla}", conexion)
    actualizar_tipos_datos(conexion, tabla, df)

    # Definir estilos para formateo condicional
    orange_fill = PatternFill(start_color="FCAF3E", end_color="FCAF3E", fill_type="solid")
    dark_orange_font = Font(color="993300")  # Naranja oscuro
    green_fill = PatternFill(start_color="ACEB67", end_color="ACEB67", fill_type="solid")   # Verde pastel
    yellow_fill = PatternFill(start_color="F5F19D", end_color="F5F19D", fill_type="solid")  # Amarillo pastel
    red_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")     # Rojo pastel

    green_font = Font(color="2B7204")   # Verde oscuro
    yellow_font = Font(color="9D8705")  # Amarillo oscuro
    red_font = Font(color="8B0000")     # Rojo oscuro

    # Obtener los índices (1-indexados) de columnas específicas que serán evaluadas en formato condicional
    col_indices = {col: df.columns.get_loc(col) + 1 for col in df.columns 
                   if col in ["Estado_TOA", "Rechazos", "Dias_Swap", "Dias_TSS", "Estado_PR_1", "Estado_PR_2", "Estado_PR_3"]}
    
    # Agregar las filas del DataFrame a la hoja de Excel (incluyendo cabecera)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            # Reemplazar valores "nan", None o 0 por cadena vacía para evitar mostrar datos no deseados
            cell_value = "" if pd.isna(value) or value in ["nan", "None"] or value is None or value == 0 else value

            # Si el valor es un Timestamp, convertirlo a datetime para Excel y aplicar formato de fecha-hora
            if isinstance(value, pd.Timestamp):
                cell_value = value.to_pydatetime()
                cell = ws.cell(row=r_idx, column=c_idx, value=cell_value)
                cell.number_format = "DD/MM/YYYY HH:MM AM/PM"
            else:
                cell = ws.cell(row=r_idx, column=c_idx, value=cell_value)

            # Si no es la primera o segunda columna, obtener las dos celdas anteriores (para formato condicional)
            if c_idx > 2:
                cell_prev = ws.cell(row=r_idx, column=c_idx-1)
                cell_prev_prev = ws.cell(row=r_idx, column=c_idx-2)

            # Aplicar formato condicional a partir de la segunda fila (omitiendo la cabecera)
            if r_idx > 1 and c_idx in col_indices.values():
                # Para columnas que deben contener valores numéricos: Rechazos, Dias_Swap y Dias_TSS
                if c_idx in [col_indices.get("Rechazos"), col_indices.get("Dias_Swap"), col_indices.get("Dias_TSS")]:
                    try:
                        numeric_value = int(cell_value)
                    except (ValueError, TypeError):
                        numeric_value = None

                # Condiciones específicas para cada columna:
                if "Rechazos" in df.columns and c_idx == col_indices.get("Rechazos") and numeric_value and numeric_value > 0:
                    cell.fill = orange_fill
                    cell.font = dark_orange_font

                elif "Dias_Swap" in df.columns and c_idx == col_indices.get("Dias_Swap") and numeric_value and numeric_value < 8:
                    cell.fill = orange_fill
                    cell.font = dark_orange_font
                    cell_prev.fill = orange_fill 
                    cell_prev.font = dark_orange_font
                    cell_prev_prev.fill = orange_fill 
                    cell_prev_prev.font = dark_orange_font

                elif "Dias_TSS" in df.columns and c_idx == col_indices.get("Dias_TSS") and numeric_value and numeric_value < 8:
                    cell.fill = orange_fill
                    cell.font = dark_orange_font
                    cell_prev.fill = orange_fill  
                    cell_prev.font = dark_orange_font

                # Para columnas de Estado_PR_x: si el valor es cadena y contiene "Pause", se formatea la celda y las dos anteriores
                elif any(f"Estado_PR_{i}" in df.columns and c_idx == col_indices.get(f"Estado_PR_{i}") 
                         and isinstance(value, str) and "Pause" in value for i in range(1, 4)):
                    cell.fill = orange_fill
                    cell.font = dark_orange_font  
                    cell_prev.fill = orange_fill 
                    cell_prev.font = dark_orange_font
                    cell_prev_prev.fill = orange_fill 
                    cell_prev_prev.font = dark_orange_font                

    # Agregar una columna con fórmula en la última columna: "Semaforo"
    ultima_columna = ws.max_column + 1
    ws.cell(row=1, column=ultima_columna, value="Semaforo")
    # Para cada fila (excepto la cabecera), si el estado no es "Completado" ni "Cancelado", se asigna una fórmula
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=col_indices.get('Estado_TOA')).value not in ["Completado", "Cancelado"]:
            # La fórmula calcula la diferencia en horas entre la fecha actual y la celda en la columna B
            ws.cell(row=row, column=ultima_columna, value=f"=(NOW()-B{row})*24").number_format = "0.00"

    # Definir el rango de la tabla en base al contenido de la hoja
    ultima_columna_letra = ws.cell(row=1, column=ultima_columna).column_letter
    rango_tabla = f"A1:{ultima_columna_letra}{ws.max_row}"

    # Crear una tabla de Excel con estilo y agregarla a la hoja
    tabla_excel = Table(displayName="TablaDatos", ref=rango_tabla)
    estilo_tabla = TableStyleInfo(name="TableStyleLight2", showRowStripes=True)
    tabla_excel.tableStyleInfo = estilo_tabla
    ws.add_table(tabla_excel)
    
    # Aplicar formato condicional a la columna "Semaforo"
    col_letter = ws.cell(row=1, column=ultima_columna).column_letter
    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{ws.max_row}",
        CellIsRule(operator="between", formula=[0.000001, 3], stopIfTrue=True, fill=green_fill, font=green_font)
    )
    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{ws.max_row}",
        CellIsRule(operator="between", formula=[3, 6], stopIfTrue=True, fill=yellow_fill, font=yellow_font)
    )
    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{ws.max_row}",
        CellIsRule(operator="greaterThan", formula=[6], stopIfTrue=True, fill=red_fill, font=red_font)
    )

    # Guardar y cerrar el archivo Excel
    wb.save(archivo_salida)
    wb.close()


def guardar_todas_las_tablas(conexion, archivo_salida):
    """
    Guarda todas las tablas de la base de datos en un único archivo Excel, 
    asignando cada tabla a una hoja independiente.

    La función realiza lo siguiente:
      - Consulta el nombre de todas las tablas existentes en la base de datos SQLite.
      - Para cada tabla, lee su contenido en un DataFrame.
      - Escribe cada DataFrame en una hoja separada del archivo Excel.
      - Se limita el nombre de la hoja a 31 caracteres (límite de Excel).

    Args:
        conexion (sqlite3.Connection): Conexión a la base de datos SQLite.
        archivo_salida (str): Ruta completa del archivo Excel de salida (por ejemplo, 'reporte_final.xlsx').
    """
    cursor = conexion.cursor()
    
    # Obtener la lista de nombres de tablas en la base de datos
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tablas = cursor.fetchall()
    
    # Crear un archivo Excel con múltiples hojas utilizando ExcelWriter de pandas
    with pd.ExcelWriter(archivo_salida, engine='openpyxl') as writer:
        for tabla in tablas:
            tabla_nombre = tabla[0]
            print(f"Guardando tabla: {tabla_nombre}")

            # Leer la tabla y almacenarla en un DataFrame
            df = pd.read_sql_query(f"SELECT * FROM {tabla_nombre}", conexion)

            # Guardar el DataFrame en una hoja separada; el nombre de la hoja se limita a 31 caracteres
            df.to_excel(writer, sheet_name=tabla_nombre[:31], index=False)


def procesar_old():
    """
    Busca las carpetas 'old' dentro de la carpeta base y sus subcarpetas,
    y elimina los archivos que tengan más de 5 días de antigüedad.
    """
    # Obtener la fecha actual
    fecha_actual = datetime.now()

    # Recorrer todas las carpetas y subcarpetas dentro de base_path
    for root, dirs, files in os.walk(base_path):
        if 'old' in dirs:
            carpeta_old = os.path.join(root, 'old')
            print(f"\tProcesando carpeta: {carpeta_old}")

            # Recorrer los archivos dentro de la carpeta 'old'
            for archivo in os.listdir(carpeta_old):
                ruta_archivo = os.path.join(carpeta_old, archivo)
                if os.path.isfile(ruta_archivo):
                    # Obtener la fecha de modificación del archivo
                    fecha_modificacion = datetime.fromtimestamp(os.path.getmtime(ruta_archivo))
                    # Calcular la diferencia en días
                    dias_antiguedad = (fecha_actual - fecha_modificacion).days
                    # Eliminar el archivo si tiene más de 5 días
                    if dias_antiguedad > 5:
                        try:
                            os.remove(ruta_archivo)
                            print(f"\tArchivo eliminado: {ruta_archivo}")
                        except OSError as e:
                            print(f"\tError al eliminar el archivo {ruta_archivo}: {e}")
    